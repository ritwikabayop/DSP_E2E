"""
Remove freeze panes from Master_Attendance.xlsx
"""
import os
from openpyxl import load_workbook

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("❄️ REMOVING FREEZE PANES - Master_Attendance.xlsx")
print("="*80)

# Open the workbook
print("\n📖 Opening Master_Attendance.xlsx...")
if not os.path.exists(MASTER_FILE):
    print(f"❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

# Process each sheet
for sheet_name in wb.sheetnames:
    print(f"\n🔧 Processing sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Check current freeze panes setting
    print(f"\n  📊 Current Freeze Panes:")
    if ws.freeze_panes:
        print(f"    Current setting: {ws.freeze_panes}")
        print(f"    Status: FROZEN ❄️")
    else:
        print(f"    Status: NOT FROZEN ✓")
    
    # Remove freeze panes
    print(f"\n  🔓 Removing freeze panes...")
    ws.freeze_panes = None
    print(f"    ✓ Freeze panes removed")
    print(f"    ✓ All columns and rows are now unfrozen")
    
    # Verify
    print(f"\n  ✅ Verification:")
    if ws.freeze_panes:
        print(f"    ❌ Still frozen at: {ws.freeze_panes}")
    else:
        print(f"    ✓ No freeze panes set - All cells are scrollable")

# Save the workbook
print(f"\n💾 Saving workbook...")
try:
    wb.save(MASTER_FILE)
    print(f"  ✓ File saved successfully!")
except Exception as e:
    print(f"  ❌ ERROR saving file: {str(e)}")
    print(f"     Make sure the file is not open in Excel")
    exit(1)

wb.close()

print("\n" + "="*80)
print("✅ FREEZE PANES REMOVED!")
print("="*80)
print("\nChanges made:")
print("  ✓ Removed freeze panes setting")
print("  ✓ Columns A, B, C are now unfrozen")
print("  ✓ All rows are now unfrozen")
print("  ✓ You can now scroll freely through all columns and rows")
print("\n📝 Next steps:")
print("  1. Close the Excel file if it's currently open")
print("  2. Reopen Master_Attendance.xlsx")
print("  3. All columns should now scroll normally without being frozen")
print("\n" + "="*80)
