"""
Delete incorrectly formatted sheets from Master_Attendance.xlsx
Keeps only January 2026 (which has the correct format)
"""
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("🗑️ CLEANING INCORRECTLY FORMATTED SHEETS")
print("="*80)

if not os.path.exists(MASTER_FILE):
    print(f"\n❌ ERROR: Master_Attendance.xlsx not found")
    exit(1)

print("\n📖 Opening Master_Attendance.xlsx...")
wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

print(f"\n📋 Current sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Keep January, delete others
sheets_to_delete = []
for sheet_name in wb.sheetnames:
    if 'january' not in sheet_name.lower():
        sheets_to_delete.append(sheet_name)

if sheets_to_delete:
    print(f"\n🗑️ Sheets to delete:")
    for sheet_name in sheets_to_delete:
        print(f"  - {sheet_name}")
    
    for sheet_name in sheets_to_delete:
        print(f"\n  Deleting '{sheet_name}'...")
        wb.remove(wb[sheet_name])
        print(f"  ✓ Deleted")
    
    # Save the workbook
    print(f"\n💾 Saving workbook...")
    try:
        wb.save(MASTER_FILE)
        print(f"  ✓ File saved successfully!")
    except Exception as e:
        print(f"  ❌ ERROR saving file: {str(e)}")
        print(f"     Make sure the file is not open in Excel")
        exit(1)
    
    print("\n" + "="*80)
    print("✅ CLEANUP COMPLETE!")
    print("="*80)
    print(f"\n✓ Deleted {len(sheets_to_delete)} sheet(s)")
    print(f"✓ Kept: Attendance January 2026 (correct format)")
    print("\n📝 Next steps:")
    print("  1. Save attendance from the web tracker for Feb/March")
    print("  2. New sheets will be created with January's format")
    print("  3. All sheets will follow the same standard format")
else:
    print(f"\n✓ No sheets to delete")
    print(f"✓ Only January 2026 exists (correct format)")

wb.close()
print("\n" + "="*80)
