"""
Extract exact format from Master_Attendance.xlsx January sheet
"""
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("📊 ANALYZING JANUARY SHEET FORMAT")
print("="*80)

wb = load_workbook(MASTER_FILE)
ws = wb['Attendance January 2026']

print(f"\n📄 Sheet: {ws.title}")
print(f"📐 Dimensions: {ws.max_row} rows × {ws.max_column} columns")

print("\n" + "="*80)
print("EXACT FORMAT - FIRST 3 ROWS")
print("="*80)

# Check first 10 columns for rows 1-3
for row in range(1, 4):
    print(f"\nRow {row}:")
    for col in range(1, 11):
        value = ws.cell(row=row, column=col).value
        print(f"  Column {col}: '{value}'")

print("\n" + "="*80)
print("SAMPLE DATA ROW (Row 4)")
print("="*80)
print(f"\nRow 4:")
for col in range(1, 11):
    value = ws.cell(row=row, column=col).value
    print(f"  Column {col}: '{value}'")

# Check if there are any merged cells
print("\n" + "="*80)
print("MERGED CELLS")
print("="*80)
if ws.merged_cells:
    print(f"\n✓ Found {len(ws.merged_cells.ranges)} merged cell ranges:")
    for merged_range in ws.merged_cells.ranges:
        print(f"  - {merged_range}")
else:
    print("\n✓ No merged cells found")

# Check column widths
print("\n" + "="*80)
print("COLUMN WIDTHS (First 10)")
print("="*80)
for col in range(1, 11):
    col_letter = ws.cell(row=1, column=col).column_letter
    width = ws.column_dimensions[col_letter].width
    print(f"  Column {col_letter}: {width}")

# Check freeze panes
print("\n" + "="*80)
print("FREEZE PANES")
print("="*80)
if ws.freeze_panes:
    print(f"✓ Freeze panes set at: {ws.freeze_panes}")
else:
    print("✓ No freeze panes")

wb.close()

print("\n" + "="*80)
print("SUMMARY OF JANUARY FORMAT")
print("="*80)
print("""
This format will be used as the base template for all future sheets.
The app.py script will be updated to match this exact structure.
""")
print("="*80)
