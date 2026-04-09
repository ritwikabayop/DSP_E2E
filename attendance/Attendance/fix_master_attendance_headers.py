"""
Fix Master_Attendance.xlsx header rows to include Lead Name column in all header rows
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("🔧 Fixing Master_Attendance.xlsx Header Rows")
print("="*80)

# Open the workbook
print("\n📖 Opening Master_Attendance.xlsx...")
if not os.path.exists(MASTER_FILE):
    print(f"❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

# Define styles
COLORS = {
    'header_bg': '0F172A',
    'header_text': 'FFFFFF',
    'subheader_bg': '334155',
    'border': '94A3B8',
}

thin_border = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='thin', color=COLORS['border']),
    bottom=Side(style='thin', color=COLORS['border'])
)

# Process each sheet
for sheet_name in wb.sheetnames:
    print(f"\n🔧 Fixing sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Check Row 1
    print(f"\n  Current Row 1:")
    print(f"    Col 1: '{ws.cell(row=1, column=1).value}'")
    print(f"    Col 2: '{ws.cell(row=1, column=2).value}'")
    print(f"    Col 3: '{ws.cell(row=1, column=3).value}'")
    
    # Fix Row 2 - Add placeholder for Lead Name column
    print(f"\n  → Fixing Row 2...")
    print(f"    Before - Col 2: '{ws.cell(row=2, column=2).value}'")
    
    # Set a placeholder or empty value for column 2 in row 2
    ws.cell(row=2, column=2).value = ""  # Empty but styled
    ws.cell(row=2, column=2).fill = PatternFill(start_color=COLORS['subheader_bg'], end_color=COLORS['subheader_bg'], fill_type='solid')
    ws.cell(row=2, column=2).font = Font(bold=True, color=COLORS['header_text'], size=9, name='Inter')
    ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=2, column=2).border = thin_border
    
    print(f"    After - Col 2: '{ws.cell(row=2, column=2).value}' (styled)")
    
    # Fix Row 3 - Add placeholder for Lead Name column
    print(f"\n  → Fixing Row 3...")
    print(f"    Before - Col 2: '{ws.cell(row=3, column=2).value}'")
    
    ws.cell(row=3, column=2).value = ""  # Empty but styled
    ws.cell(row=3, column=2).fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
    ws.cell(row=3, column=2).font = Font(bold=True, color=COLORS['header_text'], size=9, name='Inter')
    ws.cell(row=3, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=3, column=2).border = thin_border
    
    print(f"    After - Col 2: '{ws.cell(row=3, column=2).value}' (styled)")
    
    # Verify the structure
    print(f"\n  ✅ Verification:")
    print(f"    Row 1, Col 1: '{ws.cell(row=1, column=1).value}'")
    print(f"    Row 1, Col 2: '{ws.cell(row=1, column=2).value}'")
    print(f"    Row 1, Col 3: '{ws.cell(row=1, column=3).value}'")
    print(f"    Row 4, Col 1: '{ws.cell(row=4, column=1).value}'")
    print(f"    Row 4, Col 2: '{ws.cell(row=4, column=2).value}'")
    print(f"    Row 4, Col 3: '{ws.cell(row=4, column=3).value}'")

# Save the workbook
print(f"\n💾 Saving fixed workbook...")
try:
    wb.save(MASTER_FILE)
    print(f"  ✓ File saved successfully!")
except Exception as e:
    print(f"  ❌ ERROR saving file: {str(e)}")
    print(f"     Make sure the file is not open in Excel")
    exit(1)

wb.close()

print("\n" + "="*80)
print("✅ Master_Attendance.xlsx header rows fixed!")
print("="*80)
print("\nFixed issues:")
print("  ✓ Row 2, Column 2 (Lead Name) - Added styled placeholder")
print("  ✓ Row 3, Column 2 (Lead Name) - Added styled placeholder")
print("  ✓ All header rows now have consistent structure")
print("\n" + "="*80)
