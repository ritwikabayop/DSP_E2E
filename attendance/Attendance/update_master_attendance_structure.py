"""
Update Master_Attendance.xlsx structure:
1. Change first column name to "Team Member Names"
2. Add "Lead Name" column next to Team Member Names
3. Populate lead names for each team member
"""
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')
TEAM_DETAILS_FILE = os.path.join(SCRIPT_DIR, 'Team Details.xlsx')

print("="*80)
print("Updating Master_Attendance.xlsx Structure")
print("="*80)

# Step 1: Read Team Details to create member-to-lead mapping
print("\n📖 Step 1: Reading Team Details...")
if not os.path.exists(TEAM_DETAILS_FILE):
    print(f"❌ ERROR: Team Details.xlsx not found at {TEAM_DETAILS_FILE}")
    exit(1)

df_team = pd.read_excel(TEAM_DETAILS_FILE)
print(f"  ✓ Found {len(df_team)} team member records")

# Create member-to-lead mapping
member_to_lead = {}
for _, row in df_team.iterrows():
    if pd.notna(row['Team members']) and pd.notna(row['Lead']):
        member_name = str(row['Team members']).strip()
        lead_name = str(row['Lead']).strip()
        member_to_lead[member_name] = lead_name

print(f"  ✓ Created mapping for {len(member_to_lead)} team members")

# Step 2: Open Master_Attendance.xlsx
print("\n📖 Step 2: Opening Master_Attendance.xlsx...")
if not os.path.exists(MASTER_FILE):
    print(f"❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")
print(f"  ✓ Found {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}")

# Process each sheet (typically there's one sheet per month or team)
for sheet_name in wb.sheetnames:
    print(f"\n📝 Processing sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Step 3: Update first column header
    print(f"  → Step 3a: Updating first column header...")
    old_header = ws.cell(row=1, column=1).value
    print(f"     Old header: '{old_header}'")
    ws.cell(row=1, column=1).value = "Team Member Names"
    print(f"     ✓ New header: 'Team Member Names'")
    
    # Step 4: Insert new column for Lead Name (column B)
    print(f"  → Step 3b: Inserting 'Lead Name' column...")
    ws.insert_cols(2)  # Insert at column B (position 2)
    ws.cell(row=1, column=2).value = "Lead Name"
    print(f"     ✓ Inserted 'Lead Name' column at position 2")
    
    # Step 5: Populate lead names for each team member
    print(f"  → Step 3c: Populating lead names...")
    populated_count = 0
    not_found_count = 0
    not_found_members = []
    
    # Start from row 2 (row 1 is header)
    for row_idx in range(2, ws.max_row + 1):
        member_name_cell = ws.cell(row=row_idx, column=1).value
        
        if member_name_cell and str(member_name_cell).strip():
            member_name = str(member_name_cell).strip()
            
            # Look up lead name
            if member_name in member_to_lead:
                lead_name = member_to_lead[member_name]
                ws.cell(row=row_idx, column=2).value = lead_name
                populated_count += 1
            else:
                not_found_count += 1
                if len(not_found_members) < 5:  # Only show first 5
                    not_found_members.append(member_name)
    
    print(f"     ✓ Populated {populated_count} lead names")
    if not_found_count > 0:
        print(f"     ⚠ {not_found_count} team members not found in Team Details:")
        for member in not_found_members:
            print(f"        - {member}")
        if not_found_count > 5:
            print(f"        ... and {not_found_count - 5} more")

# Step 6: Save the workbook
print(f"\n💾 Step 4: Saving changes to Master_Attendance.xlsx...")
try:
    wb.save(MASTER_FILE)
    print(f"  ✓ File saved successfully!")
except Exception as e:
    print(f"  ❌ ERROR saving file: {str(e)}")
    print(f"     Make sure the file is not open in Excel")
    exit(1)

wb.close()

print("\n" + "="*80)
print("✅ Master_Attendance.xlsx structure updated successfully!")
print("="*80)
print("\nChanges made:")
print("  1. ✓ First column renamed to 'Team Member Names'")
print("  2. ✓ Added 'Lead Name' column at position 2")
print("  3. ✓ Populated lead names for all team members")
print("\n" + "="*80)
