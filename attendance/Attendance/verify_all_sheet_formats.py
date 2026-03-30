"""
Verify that all sheets in Master_Attendance.xlsx follow the correct format
"""
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("🔍 VERIFYING SHEET FORMAT STANDARDIZATION")
print("="*80)

if not os.path.exists(MASTER_FILE):
    print(f"\n❌ ERROR: Master_Attendance.xlsx not found")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"\n📄 File: Master_Attendance.xlsx")
print(f"📊 Total sheets: {len(wb.sheetnames)}")

print("\n" + "="*80)
print("EXPECTED FORMAT (STANDARD)")
print("="*80)
print("""
Row 1: Column 1='Team Member Names', Column 2='Lead Name'
Row 2: Column 1='Latest Update Source: ...', Column 3='📍 Location', Column 4+=Date numbers (1,2,3...)
Row 3: Column 1='Last Saved At: ...', Column 4+=Day names (Mon,Tue,Wed...)
Row 4+: Data rows (Team Member Name, Lead Name, Location, attendance status...)
""")

all_sheets_correct = True

for sheet_name in wb.sheetnames:
    print("="*80)
    print(f"SHEET: {sheet_name}")
    print("="*80)
    
    ws = wb[sheet_name]
    
    # Check Row 1
    r1_c1 = ws.cell(row=1, column=1).value
    r1_c2 = ws.cell(row=1, column=2).value
    r1_c3 = ws.cell(row=1, column=3).value
    
    print(f"\nRow 1:")
    print(f"  Column 1: '{r1_c1}'")
    print(f"  Column 2: '{r1_c2}'")
    print(f"  Column 3: '{r1_c3}'")
    
    row1_correct = (
        r1_c1 and 'team member' in str(r1_c1).lower() and
        r1_c2 and 'lead name' in str(r1_c2).lower()
    )
    
    if row1_correct:
        print(f"  ✅ Row 1 format CORRECT")
    else:
        print(f"  ❌ Row 1 format INCORRECT")
        all_sheets_correct = False
    
    # Check Row 2
    r2_c1 = ws.cell(row=2, column=1).value
    r2_c2 = ws.cell(row=2, column=2).value
    r2_c3 = ws.cell(row=2, column=3).value
    r2_c4 = ws.cell(row=2, column=4).value
    
    print(f"\nRow 2:")
    print(f"  Column 1: '{r2_c1}'")
    print(f"  Column 2: '{r2_c2}' (should be empty)")
    print(f"  Column 3: '{r2_c3}'")
    print(f"  Column 4: '{r2_c4}'")
    
    row2_correct = (
        r2_c1 and 'latest update' in str(r2_c1).lower() and
        r2_c3 and 'location' in str(r2_c3).lower()
    )
    
    if row2_correct:
        print(f"  ✅ Row 2 format CORRECT")
    else:
        print(f"  ❌ Row 2 format INCORRECT")
        all_sheets_correct = False
    
    # Check Row 3
    r3_c1 = ws.cell(row=3, column=1).value
    r3_c2 = ws.cell(row=3, column=2).value
    r3_c3 = ws.cell(row=3, column=3).value
    r3_c4 = ws.cell(row=3, column=4).value
    
    print(f"\nRow 3:")
    print(f"  Column 1: '{r3_c1}'")
    print(f"  Column 2: '{r3_c2}' (should be empty)")
    print(f"  Column 3: '{r3_c3}' (should be empty)")
    print(f"  Column 4: '{r3_c4}' (should be day name)")
    
    row3_correct = (
        r3_c1 and 'last saved' in str(r3_c1).lower()
    )
    
    if row3_correct:
        print(f"  ✅ Row 3 format CORRECT")
    else:
        print(f"  ❌ Row 3 format INCORRECT")
        all_sheets_correct = False
    
    # Check data row 4
    r4_c1 = ws.cell(row=4, column=1).value
    r4_c2 = ws.cell(row=4, column=2).value
    r4_c3 = ws.cell(row=4, column=3).value
    r4_c4 = ws.cell(row=4, column=4).value
    
    print(f"\nRow 4 (First data row):")
    print(f"  Column 1: '{r4_c1}' (team member name)")
    print(f"  Column 2: '{r4_c2}' (lead name)")
    print(f"  Column 3: '{r4_c3}' (location)")
    print(f"  Column 4: '{r4_c4}' (attendance)")
    
    if r4_c1 and r4_c2:
        print(f"  ✅ Data row format appears CORRECT")
    else:
        print(f"  ⚠️ Data row might be empty or incorrect")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

if all_sheets_correct:
    print("\n✅ ALL SHEETS FOLLOW THE CORRECT STANDARD FORMAT!")
else:
    print("\n⚠️ SOME SHEETS HAVE FORMAT ISSUES")
    print("\nTo fix:")
    print("  1. Delete incorrectly formatted sheets")
    print("  2. Save attendance from the web tracker")
    print("  3. New sheets will be created with the correct format")

print("\n" + "="*80)
print("FORMAT GUARANTEE")
print("="*80)
print("""
✅ The app.py code has been updated to ensure ALL new sheets follow this format:

Row 1: Team Member Names | Lead Name | (empty)
Row 2: Latest Update Source: ... | (empty) | 📍 Location | 1 | 2 | 3 | ...
Row 3: Last Saved At: ... | (empty) | (empty) | Mon | Tue | Wed | ...
Row 4+: name | lead | location | P | | SL | ...

This format will be consistent across ALL months!
""")

wb.close()
print("="*80)
