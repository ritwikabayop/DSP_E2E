"""
Make Master_Attendance.xlsx colorful with professional styling
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("🎨 Making Master_Attendance.xlsx Colorful")
print("="*80)

# Open the workbook
print("\n📖 Opening Master_Attendance.xlsx...")
if not os.path.exists(MASTER_FILE):
    print(f"❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

# Define color schemes
COLORS = {
    'header_bg': '0F172A',        # Dark slate for main headers
    'header_text': 'FFFFFF',       # White text
    'subheader_bg': '334155',      # Medium slate for subheaders
    'team_member_bg': 'F1F5F9',    # Light slate for team member column
    'lead_name_bg': 'DBEAFE',      # Light blue for lead name column
    'location_bg': 'FEF3C7',       # Light yellow for location
    'present_bg': 'BBF7D0',        # Green for present
    'leave_bg': 'FEE2E2',          # Red for leaves
    'weekend_bg': 'E0E7FF',        # Light blue for weekends
    'holiday_bg': 'FEF9C3',        # Yellow for holidays
    'border': '94A3B8',            # Slate for borders
    'row_even': 'F8FAFC',          # Very light gray for even rows
    'row_odd': 'FFFFFF',           # White for odd rows
}

# Define styles
thin_border = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='thin', color=COLORS['border']),
    bottom=Side(style='thin', color=COLORS['border'])
)

thick_border = Border(
    left=Side(style='medium', color=COLORS['border']),
    right=Side(style='medium', color=COLORS['border']),
    top=Side(style='medium', color=COLORS['border']),
    bottom=Side(style='medium', color=COLORS['border'])
)

# Process each sheet
for sheet_name in wb.sheetnames:
    print(f"\n🎨 Styling sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Style the main header row (Row 1)
    print(f"  → Styling header row...")
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        
        # Different colors for different columns
        if col_idx == 1:  # Team Member Names
            cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        elif col_idx == 2:  # Lead Name
            cell.fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')
        elif col_idx == 3:  # Location
            cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
        else:  # Date columns
            cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
        
        cell.font = Font(bold=True, color=COLORS['header_text'], size=11, name='Inter')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thick_border
    
    # Style row 2 (metadata row with day names)
    print(f"  → Styling day names row...")
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = PatternFill(start_color=COLORS['subheader_bg'], end_color=COLORS['subheader_bg'], fill_type='solid')
        cell.font = Font(bold=True, color=COLORS['header_text'], size=9, name='Inter')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Style row 3 (weekday row)
    print(f"  → Styling weekday row...")
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = PatternFill(start_color='475569', end_color='475569', fill_type='solid')
        cell.font = Font(bold=True, color=COLORS['header_text'], size=9, name='Inter')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Style data rows (starting from row 4)
    print(f"  → Styling data rows...")
    styled_rows = 0
    for row_idx in range(4, ws.max_row + 1):
        # Alternate row colors
        is_even = (row_idx % 2 == 0)
        row_bg = COLORS['row_even'] if is_even else COLORS['row_odd']
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Column-specific styling
            if col_idx == 1:  # Team Member Names
                cell.fill = PatternFill(start_color=COLORS['team_member_bg'], end_color=COLORS['team_member_bg'], fill_type='solid')
                cell.font = Font(bold=True, color='0F172A', size=10, name='Inter')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='medium', color='3B82F6'),
                    right=Side(style='thin', color=COLORS['border']),
                    top=Side(style='thin', color=COLORS['border']),
                    bottom=Side(style='thin', color=COLORS['border'])
                )
            
            elif col_idx == 2:  # Lead Name
                cell.fill = PatternFill(start_color=COLORS['lead_name_bg'], end_color=COLORS['lead_name_bg'], fill_type='solid')
                cell.font = Font(bold=False, color='1E40AF', size=10, name='Inter')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
            
            elif col_idx == 3:  # Location
                cell.fill = PatternFill(start_color=COLORS['location_bg'], end_color=COLORS['location_bg'], fill_type='solid')
                cell.font = Font(bold=False, color='92400E', size=9, name='Inter')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color=COLORS['border']),
                    right=Side(style='medium', color='94A3B8'),
                    top=Side(style='thin', color=COLORS['border']),
                    bottom=Side(style='thin', color=COLORS['border'])
                )
            
            else:  # Attendance data columns
                cell_value = str(cell.value).upper() if cell.value else ''
                
                # Apply colors based on attendance status
                if cell_value == 'P':
                    cell.fill = PatternFill(start_color=COLORS['present_bg'], end_color=COLORS['present_bg'], fill_type='solid')
                    cell.font = Font(bold=True, color='166534', size=10)
                elif cell_value in ['VG', 'VR', 'C', 'SL', 'UP', 'CL', 'CG', 'ML']:
                    cell.fill = PatternFill(start_color=COLORS['leave_bg'], end_color=COLORS['leave_bg'], fill_type='solid')
                    cell.font = Font(bold=True, color='991B1B', size=10)
                elif cell_value == 'MH':
                    cell.fill = PatternFill(start_color=COLORS['holiday_bg'], end_color=COLORS['holiday_bg'], fill_type='solid')
                    cell.font = Font(bold=True, color='854D0E', size=10)
                elif cell_value in ['T', 'AT']:
                    cell.fill = PatternFill(start_color='DDD6FE', end_color='DDD6FE', fill_type='solid')
                    cell.font = Font(bold=True, color='5B21B6', size=10)
                elif cell_value == 'OP':
                    cell.fill = PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid')
                    cell.font = Font(bold=True, color='92400E', size=10)
                elif cell_value == 'W':
                    cell.fill = PatternFill(start_color=COLORS['present_bg'], end_color=COLORS['present_bg'], fill_type='solid')
                    cell.font = Font(bold=True, color='166534', size=10)
                else:
                    # Check if it's a weekend by looking at row 3
                    day_name = str(ws.cell(row=3, column=col_idx).value).upper() if ws.cell(row=3, column=col_idx).value else ''
                    if 'SAT' in day_name or 'SUN' in day_name:
                        cell.fill = PatternFill(start_color=COLORS['weekend_bg'], end_color=COLORS['weekend_bg'], fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type='solid')
                    cell.font = Font(size=10)
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        styled_rows += 1
        if styled_rows % 20 == 0:
            print(f"     Styled {styled_rows} rows...")
    
    print(f"  ✓ Styled {styled_rows} data rows")
    
    # Adjust column widths
    print(f"  → Adjusting column widths...")
    ws.column_dimensions['A'].width = 25  # Team Member Names
    ws.column_dimensions['B'].width = 20  # Lead Name
    ws.column_dimensions['C'].width = 22  # Location
    
    # Set width for date columns
    for col_idx in range(4, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 6
    
    print(f"  ✓ Column widths adjusted")
    
    # Freeze panes (freeze first 3 rows and first 3 columns)
    print(f"  → Freezing panes...")
    ws.freeze_panes = 'D4'  # Freeze everything before column D and row 4
    print(f"  ✓ Panes frozen at D4")

# Save the workbook
print(f"\n💾 Saving colorful workbook...")
try:
    wb.save(MASTER_FILE)
    print(f"  ✓ File saved successfully!")
except Exception as e:
    print(f"  ❌ ERROR saving file: {str(e)}")
    print(f"     Make sure the file is not open in Excel")
    exit(1)

wb.close()

print("\n" + "="*80)
print("✅ Master_Attendance.xlsx is now colorful and styled!")
print("="*80)
print("\n🎨 Applied styling:")
print("  ✓ Dark slate headers with white text")
print("  ✓ Blue highlight for Team Member Names column")
print("  ✓ Purple highlight for Lead Name column")  
print("  ✓ Amber highlight for Location column")
print("  ✓ Green background for Present (P)")
print("  ✓ Red background for Leaves (VG, VR, SL, etc.)")
print("  ✓ Yellow background for Holidays (MH, OP)")
print("  ✓ Purple background for Training (T, AT)")
print("  ✓ Light blue for weekends")
print("  ✓ Alternating row colors for better readability")
print("  ✓ Professional borders and spacing")
print("  ✓ Frozen panes for easy navigation")
print("\n" + "="*80)
