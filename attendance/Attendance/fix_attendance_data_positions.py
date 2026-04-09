"""
Fix attendance data column positions in Master_Attendance.xlsx
This script shifts attendance data from the wrong columns to the correct columns
after the Lead Name column was added
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

# Cell background colors matching the HTML tracker
STATUS_COLORS = {
    'P':  'C8F7C5',  # soft mint green
    'VG': 'B3DDF2',  # soft sky blue
    'VR': 'D4EBF2',  # soft powder blue
    'C':  '9EE5DB',  # soft turquoise
    'SL': 'FFB3B3',  # soft red
    'OP': 'FFE69C',  # soft gold
    'MH': 'FFF9C4',  # soft yellow
    'UP': 'FFB3D9',  # soft pink
    'CL': 'D4C5F9',  # soft lavender
    'CG': 'FFB3D1',  # soft rose
    'T':  'E8D5F5',  # soft orchid
    'AT': 'C5B3E6',  # soft purple
    'ML': 'F0D5F0',  # soft plum
}

def get_fill(status):
    """Return a PatternFill for the given status, or empty for blank/unknown."""
    code = (status or '').strip().upper()
    if code in STATUS_COLORS:
        return PatternFill(fill_type='solid', fgColor=STATUS_COLORS[code])
    return PatternFill(fill_type=None)

print("="*80)
print("🔧 FIXING ATTENDANCE DATA COLUMN POSITIONS")
print("="*80)

# Open the workbook
print("\n📖 Opening Master_Attendance.xlsx...")
if not os.path.exists(MASTER_FILE):
    print(f"❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

# Process each sheet
for sheet_name in wb.sheetnames:
    print(f"\n🔧 Processing sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Check header structure
    print(f"\n  📊 Current Header Structure:")
    header_row_1 = ws.cell(row=1, column=1).value
    header_row_1_col2 = ws.cell(row=1, column=2).value
    header_row_1_col3 = ws.cell(row=1, column=3).value
    
    print(f"    Row 1, Col 1: '{header_row_1}'")
    print(f"    Row 1, Col 2: '{header_row_1_col2}'")
    print(f"    Row 1, Col 3: '{header_row_1_col3}'")
    
    # Look for the structure - we already have the Lead Name column at position 2
    # But attendance data might be in positions starting at column 3 when it should start at column 4
    
    # Check if row 1 has the header
    if header_row_1_col2 == 'Lead Name':
        print(f"\n  ✅ Lead Name column exists at position 2")
        print(f"  ℹ️ Checking if attendance data needs to be shifted...")
        
        # Find the actual data start row (usually row 4)
        data_start_row = 4
        
        # Check if column 3 incorrectly has attendance data instead of just location
        # The location column should have text like "Bengaluru, BDC6F"
        # Date columns should have status codes like "P", "SL", etc.
        
        needs_fix = False
        sample_row = data_start_row
        location_value = ws.cell(row=sample_row, column=3).value
        
        # Check if column 3 (location) contains a status code (means data is misaligned)
        if location_value and str(location_value).strip().upper() in STATUS_COLORS:
            needs_fix = True
            print(f"  ⚠️ Column 3 contains status code '{location_value}' - data is misaligned!")
        else:
            print(f"  ✓ Column 3 contains location info - data appears correctly aligned")
            continue
        
        if not needs_fix:
            print(f"  ✓ Data already correctly positioned, skipping this sheet")
            continue
        
        print(f"\n  🔄 Shifting attendance data from columns 3+ to columns 4+...")
        
        # Determine the number of date columns (usually 31 for a month)
        # We'll move ALL data from row 4 onwards, columns 3 to end, shifted right by 1
        
        total_rows_processed = 0
        
        # Process from bottom to top to avoid overwriting
        # Start from max_row and go down to data_start_row
        for row_idx in range(ws.max_row, data_start_row - 1, -1):
            # Check if this is a data row (has a name in column 1)
            name_cell = ws.cell(row=row_idx, column=1).value
            if not name_cell or not str(name_cell).strip():
                continue
            
            # Skip metadata rows
            name_str = str(name_cell).strip().lower()
            if 'team member' in name_str or 'latest update' in name_str or 'last saved' in name_str:
                continue
            
            # Get the last column with data (need to determine max column for this row)
            last_col = ws.max_column
            
            # Move data from right to left to avoid overwriting
            for col_idx in range(last_col, 2, -1):  # From last column down to column 3
                source_col = col_idx - 1  # Read from one column to the left
                target_col = col_idx      # Write to current column
                
                # Skip if source is column 2 (Lead Name) or column 1 (Team Member)
                if source_col <= 2:
                    continue
                
                # Copy cell value and style
                source_cell = ws.cell(row=row_idx, column=source_col)
                target_cell = ws.cell(row=row_idx, column=target_col)
                
                target_cell.value = source_cell.value
                
                # Apply status color if it's a status code
                if source_cell.value:
                    target_cell.fill = get_fill(str(source_cell.value))
                else:
                    target_cell.fill = PatternFill(fill_type=None)
            
            # Clear column 3 (Location) as it should have been text, not status
            # We'll need to restore location from somewhere or leave it for now
            ws.cell(row=row_idx, column=3).value = None
            ws.cell(row=row_idx, column=3).fill = PatternFill(fill_type=None)
            
            total_rows_processed += 1
            
            if total_rows_processed % 20 == 0:
                print(f"    Processed {total_rows_processed} rows...")
        
        print(f"  ✅ Shifted {total_rows_processed} rows")
        print(f"  ⚠️ Note: Location data in column 3 has been cleared")
        print(f"      You may need to re-save attendance from the web tracker to restore locations")
    
    else:
        print(f"  ℹ️ Structure doesn't match expected format, skipping")

# Save the workbook
print(f"\n💾 Saving fixed workbook...")
try:
    wb.save(MASTER_FILE)
    print(f"  ✓ File saved successfully!")
except Exception as e:
    print(f"  ❌ ERROR saving file: {str(e)}")
    print(f"     Make sure the file is not open in Excel")
    exit(1)

wb.close()

print("\n" + "="*80)
print("✅ ATTENDANCE DATA COLUMN POSITIONS FIXED!")
print("="*80)
print("\nWhat was fixed:")
print("  ✓ Attendance data shifted from column 3+ to column 4+")
print("  ✓ Column structure now matches:")
print("     - Column 1: Team Member Names")
print("  ✓ Column 2: Lead Name")
print("     - Column 3: Location (will be populated on next save)")
print("     - Column 4+: Date columns with attendance data")
print("\n📝 Next steps:")
print("  1. Close Excel if it's open")
print("  2. Reopen Master_Attendance.xlsx to verify")
print("  3. Save attendance from the web tracker once to restore location data")
print("\n" + "="*80)
