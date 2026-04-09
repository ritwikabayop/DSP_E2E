"""
Delete incorrectly formatted February 2026 sheet from Master_Attendance.xlsx
"""
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("🗑️ DELETING INCORRECTLY FORMATTED FEBRUARY SHEET")
print("="*80)

if not os.path.exists(MASTER_FILE):
    print(f"\n❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

print("\n📖 Opening Master_Attendance.xlsx...")
wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

print(f"\n📋 Current sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Look for February 2026 sheet
feb_sheet_name = None
for sheet_name in wb.sheetnames:
    if 'february' in sheet_name.lower() and '2026' in sheet_name:
        feb_sheet_name = sheet_name
        break

if feb_sheet_name:
    print(f"\n🗑️ Deleting sheet: '{feb_sheet_name}'")
    wb.remove(wb[feb_sheet_name])
    print(f"  ✓ Sheet deleted")
    
    # Save the workbook
    print(f"\n💾 Saving workbook...")
    try:
        wb.save(MASTER_FILE)
        print(f"  ✓ File saved successfully!")
    except Exception as e:
        print(f"  ❌ ERROR saving file: {str(e)}")
        print(f"     Make sure the file is not open in Excel")
        exit(1)
    
    print("\n" + "="*80)
    print("✅ FEBRUARY SHEET DELETED!")
    print("="*80)
    print("\nNext steps:")
    print("  1. Save attendance from the web tracker for February")
    print("  2. The new sheet will be created with the correct format")
    print("  3. It will match January 2026's format exactly")
else:
    print(f"\n⚠️ No February 2026 sheet found")
    print(f"   Available sheets: {', '.join(wb.sheetnames)}")

wb.close()
print("\n" + "="*80)
