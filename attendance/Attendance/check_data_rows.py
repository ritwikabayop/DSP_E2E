"""
Get the exact data from row 4 onwards
"""
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

wb = load_workbook(MASTER_FILE)
ws = wb['Attendance January 2026']

print("="*80)
print("ROW 4 AND BEYOND (DATA ROWS)")
print("="*80)

for row in range(4, 8):  # Check rows 4-7
    print(f"\nRow {row}:")
    print(f"  Col 1 (Team Member): '{ws.cell(row=row, column=1).value}'")
    print(f"  Col 2 (Lead Name): '{ws.cell(row=row, column=2).value}'")
    print(f"  Col 3 (Location): '{ws.cell(row=row, column=3).value}'")
    print(f"  Col 4 (Day 1): '{ws.cell(row=row, column=4).value}'")
    print(f"  Col 5 (Day 2): '{ws.cell(row=row, column=5).value}'")

wb.close()
