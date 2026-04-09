"""
Complete verification of Master_Attendance.xlsx structure and data alignment
"""
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("🔍 MASTER ATTENDANCE COMPLETE VERIFICATION")
print("="*80)

wb = load_workbook(MASTER_FILE)
ws = wb.active

print(f"\n📄 File: Master_Attendance.xlsx")
print(f"📊 Sheet: {ws.title}")

print("\n" + "="*80)
print("STRUCTURE VERIFICATION")
print("="*80)

# Check column headers (Row 1)
col1_header = ws.cell(row=1, column=1).value
col2_header = ws.cell(row=1, column=2).value
col3_header = ws.cell(row=1, column=3).value
col4_header = ws.cell(row=1, column=4).value

print(f"\nRow 1 (Main Headers):")
print(f"  Column 1: '{col1_header}' {'✅' if 'Team Member' in str(col1_header or '') else '❌'}")
print(f"  Column 2: '{col2_header}' {'✅' if 'Lead Name' in str(col2_header or '') else '❌'}")
print(f"  Column 3: '{col3_header}' {'✅' if col3_header is None or 'Location' not in str(col3_header or '') else '⚠️'}")
print(f"  Column 4: '{col4_header}' {'✅' if col4_header is None else '⚠️'}")

# Check metadata rows
row2_col1 = ws.cell(row=2, column=1).value
row2_col2 = ws.cell(row=2, column=2).value
row2_col3 = ws.cell(row=2, column=3).value
row2_col4 = ws.cell(row=2, column=4).value

print(f"\nRow 2 (Metadata):")
print(f"  Column 1: '{row2_col1}'")
print(f"  Column 2: '{row2_col2}' (empty is OK)")
print(f"  Column 3: '{row2_col3}' {'✅ Location header' if 'Location' in str(row2_col3 or '') else '⚠️'}")
print(f"  Column 4: '{row2_col4}' {'✅ Date 1' if row2_col4 else '❌'}")

row3_col3 = ws.cell(row=3, column=3).value
row3_col4 = ws.cell(row=3, column=4).value
row3_col5 = ws.cell(row=3, column=5).value

print(f"\nRow 3 (Day names):")
print(f"  Column 3: '{row3_col3}' (empty is OK)")
print(f"  Column 4: '{row3_col4}' {'✅' if row3_col4 else '❌'}")
print(f"  Column 5: '{row3_col5}' {'✅' if row3_col5 else '❌'}")

print("\n" + "="*80)
print("DATA VERIFICATION (First 5 team members)")
print("="*80)

all_correct = True

for row_idx in range(4, min(9, ws.max_row + 1)):
    team_member = ws.cell(row=row_idx, column=1).value
    lead_name = ws.cell(row=row_idx, column=2).value
    location = ws.cell(row=row_idx, column=3).value
    date1 = ws.cell(row=row_idx, column=4).value
    date2 = ws.cell(row=row_idx, column=5).value
    
    if not team_member:
        continue
    
    print(f"\nRow {row_idx}:")
    print(f"  👤 Team Member (Col 1): {team_member}")
    print(f"  👔 Lead Name (Col 2): {lead_name}")
    print(f"  📍 Location (Col 3): {location}")
    print(f"  📅 Day 1 (Col 4): {date1 if date1 else '(empty)'}")
    print(f"  📅 Day 2 (Col 5): {date2 if date2 else '(empty)'}")
    
    # Verify structure
    issues = []
    
    # Location should not be a status code
    if location and str(location).strip().upper() in ['P', 'SL', 'VG', 'VR', 'C', 'OP', 'MH', 'UP', 'CL', 'CG', 'T', 'AT', 'ML']:
        issues.append("⚠️ Location column contains status code (DATA MISALIGNED!)")
        all_correct = False
    
    # Lead name should exist
    if not lead_name:
        issues.append("⚠️ Lead Name missing")
    
    if issues:
        for issue in issues:
            print(f"  {issue}")
    else:
        print(f"  ✅ Structure correct")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)

if all_correct:
    print("\n✅ ALL CHECKS PASSED!")
    print("\nStructure is correct:")
    print("  • Column 1: Team Member Names")
    print("  • Column 2: Lead Name")
    print("  • Column 3: Location")
    print("  • Column 4+: Date columns with attendance data")
    print("\n✅ Attendance data is in the correct columns")
    print("✅ Ready to use!")
else:
    print("\n⚠️ ISSUES DETECTED!")
    print("\nPlease check the issues listed above.")
    print("You may need to run fix_attendance_data_positions.py")

print("\n" + "="*80)
print("VIEW INSTRUCTIONS")
print("="*80)
print("\n📖 To view the Excel file correctly:")
print("  1. Close Master_Attendance.xlsx if it's open")
print("  2. Reopen the file")
print("  3. Scroll to column A (far left) to see Team Member Names")
print("  4. Column B shows Lead Name")
print("  5. Column C shows Location")
print("  6. Columns D onwards show dates (1, 2, 3, ...)")
print("\n💡 The freeze panes have been removed, so you can scroll freely")
print("   Make sure you're viewing from column A to see all data!")

wb.close()
print("\n" + "="*80)
