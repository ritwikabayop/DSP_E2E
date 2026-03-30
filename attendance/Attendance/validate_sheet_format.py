"""
Validate Master_Attendance.xlsx sheets against the standard format
This script checks if all sheets follow the January 2026 format specification
"""
import os
from openpyxl import load_workbook
import calendar

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

def validate_sheet(ws, sheet_name):
    """Validate a single sheet against the standard format"""
    print(f"\n{'='*80}")
    print(f"VALIDATING: {sheet_name}")
    print(f"{'='*80}")
    
    errors = []
    warnings = []
    
    # Check Row 1
    r1c1 = ws.cell(row=1, column=1).value
    r1c2 = ws.cell(row=1, column=2).value
    
    print(f"\n✓ Checking Row 1 (Main Headers)...")
    if not r1c1 or 'team member' not in str(r1c1).lower():
        errors.append("Row 1, Column 1 should be 'Team Member Names'")
        print(f"  ❌ Column 1: '{r1c1}' (expected 'Team Member Names')")
    else:
        print(f"  ✓ Column 1: '{r1c1}'")
    
    if not r1c2 or 'lead name' not in str(r1c2).lower():
        errors.append("Row 1, Column 2 should be 'Lead Name'")
        print(f"  ❌ Column 2: '{r1c2}' (expected 'Lead Name')")
    else:
        print(f"  ✓ Column 2: '{r1c2}'")
    
    # Check Row 2
    r2c1 = ws.cell(row=2, column=1).value
    r2c3 = ws.cell(row=2, column=3).value
    r2c4 = ws.cell(row=2, column=4).value
    
    print(f"\n✓ Checking Row 2 (Metadata + Date Headers)...")
    if not r2c1 or 'latest update' not in str(r2c1).lower():
        errors.append("Row 2, Column 1 should start with 'Latest Update Source:'")
        print(f"  ❌ Column 1: '{r2c1}' (expected 'Latest Update Source: ...')")
    else:
        print(f"  ✓ Column 1: '{r2c1}'")
    
    if not r2c3 or 'location' not in str(r2c3).lower():
        errors.append("Row 2, Column 3 should be '📍 Location'")
        print(f"  ❌ Column 3: '{r2c3}' (expected '📍 Location')")
    else:
        print(f"  ✓ Column 3: '{r2c3}'")
    
    if r2c4 != 1 and str(r2c4) != '1':
        errors.append("Row 2, Column 4 should be '1' (first date)")
        print(f"  ❌ Column 4: '{r2c4}' (expected '1')")
    else:
        print(f"  ✓ Column 4: '{r2c4}' (first date)")
    
    # Check Row 3
    r3c1 = ws.cell(row=3, column=1).value
    r3c4 = ws.cell(row=3, column=4).value
    
    print(f"\n✓ Checking Row 3 (Timestamp + Day Names)...")
    if not r3c1 or 'last saved' not in str(r3c1).lower():
        errors.append("Row 3, Column 1 should start with 'Last Saved At:'")
        print(f"  ❌ Column 1: '{r3c1}' (expected 'Last Saved At: ...')")
    else:
        print(f"  ✓ Column 1: '{r3c1}'")
    
    if not r3c4 or len(str(r3c4)) < 2:
        warnings.append("Row 3, Column 4 should have day name (e.g., 'Mon', 'Thu')")
        print(f"  ⚠️ Column 4: '{r3c4}' (expected day name like 'Mon')")
    else:
        print(f"  ✓ Column 4: '{r3c4}' (day name)")
    
    # Check Row 4 (First data row)
    r4c1 = ws.cell(row=4, column=1).value
    r4c2 = ws.cell(row=4, column=2).value
    r4c3 = ws.cell(row=4, column=3).value
    
    print(f"\n✓ Checking Row 4 (First Data Row)...")
    if not r4c1 or str(r4c1).strip() == '':
        warnings.append("Row 4, Column 1 should have team member name")
        print(f"  ⚠️ Column 1: '{r4c1}' (expected team member name)")
    else:
        print(f"  ✓ Column 1: '{r4c1}' (team member)")
    
    if not r4c2 or str(r4c2).strip() == '':
        warnings.append("Row 4, Column 2 should have lead name")
        print(f"  ⚠️ Column 2: '{r4c2}' (expected lead name)")
    else:
        print(f"  ✓ Column 2: '{r4c2}' (lead name)")
    
    if not r4c3 or str(r4c3).strip() == '':
        warnings.append("Row 4, Column 3 should have location")
        print(f"  ⚠️ Column 3: '{r4c3}' (expected location)")
    else:
        print(f"  ✓ Column 3: '{r4c3}' (location)")
    
    # Check column widths
    print(f"\n✓ Checking Column Widths...")
    col_a_width = ws.column_dimensions['A'].width
    col_b_width = ws.column_dimensions['B'].width
    col_c_width = ws.column_dimensions['C'].width
    
    if abs(col_a_width - 30.0) > 1:
        warnings.append(f"Column A width is {col_a_width}, expected 30.0")
        print(f"  ⚠️ Column A: {col_a_width} (expected 30.0)")
    else:
        print(f"  ✓ Column A: {col_a_width}")
    
    if abs(col_b_width - 25.0) > 1:
        warnings.append(f"Column B width is {col_b_width}, expected 25.0")
        print(f"  ⚠️ Column B: {col_b_width} (expected 25.0)")
    else:
        print(f"  ✓ Column B: {col_b_width}")
    
    if abs(col_c_width - 22.0) > 1:
        warnings.append(f"Column C width is {col_c_width}, expected 22.0")
        print(f"  ⚠️ Column C: {col_c_width} (expected 22.0)")
    else:
        print(f"  ✓ Column C: {col_c_width}")
    
    # Check for merged cells
    print(f"\n✓ Checking Merged Cells...")
    if ws.merged_cells:
        warnings.append(f"Sheet has {len(ws.merged_cells.ranges)} merged cell ranges (expected 0)")
        print(f"  ⚠️ Found {len(ws.merged_cells.ranges)} merged cell ranges")
    else:
        print(f"  ✓ No merged cells")
    
    # Check freeze panes
    print(f"\n✓ Checking Freeze Panes...")
    if ws.freeze_panes:
        warnings.append(f"Sheet has freeze panes at {ws.freeze_panes} (expected None)")
        print(f"  ⚠️ Freeze panes set at: {ws.freeze_panes}")
    else:
        print(f"  ✓ No freeze panes")
    
    # Summary
    print(f"\n{'='*80}")
    print(f"SUMMARY FOR {sheet_name}")
    print(f"{'='*80}")
    
    if errors:
        print(f"\n❌ ERRORS ({len(errors)}):")
        for i, error in enumerate(errors, 1):
            print(f"  {i}. {error}")
    
    if warnings:
        print(f"\n⚠️  WARNINGS ({len(warnings)}):")
        for i, warning in enumerate(warnings, 1):
            print(f"  {i}. {warning}")
    
    if not errors and not warnings:
        print(f"\n✅ ALL CHECKS PASSED!")
        print(f"   {sheet_name} follows the standard format perfectly.")
    elif not errors:
        print(f"\n✅ NO CRITICAL ERRORS")
        print(f"   {sheet_name} has minor format issues (see warnings).")
    else:
        print(f"\n❌ VALIDATION FAILED")
        print(f"   {sheet_name} has critical format errors.")
    
    return len(errors) == 0, len(warnings) == 0

def main():
    print("="*80)
    print("MASTER ATTENDANCE FORMAT VALIDATOR")
    print("="*80)
    print(f"\n📄 File: {MASTER_FILE}")
    
    if not os.path.exists(MASTER_FILE):
        print(f"\n❌ ERROR: File not found!")
        return False
    
    wb = load_workbook(MASTER_FILE)
    print(f"✓ Workbook loaded: {len(wb.sheetnames)} sheet(s)")
    
    all_passed = True
    all_perfect = True
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        passed, perfect = validate_sheet(ws, sheet_name)
        all_passed = all_passed and passed
        all_perfect = all_perfect and perfect
    
    wb.close()
    
    # Final summary
    print(f"\n{'='*80}")
    print("FINAL SUMMARY")
    print(f"{'='*80}")
    print(f"\nTotal sheets validated: {len(wb.sheetnames)}")
    
    if all_perfect:
        print(f"\n✅ ALL SHEETS PERFECT!")
        print("   All sheets follow the standard January 2026 format exactly.")
    elif all_passed:
        print(f"\n✅ ALL SHEETS PASSED")
        print("   All sheets follow the format with minor warnings.")
    else:
        print(f"\n❌ SOME SHEETS FAILED")
        print("   Some sheets have critical format errors.")
    
    print(f"\n{'='*80}")
    return all_passed

if __name__ == "__main__":
    success = main()
    print()
    exit(0 if success else 1)
