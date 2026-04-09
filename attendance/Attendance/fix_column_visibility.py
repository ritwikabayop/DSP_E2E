"""
Fix column visibility for Team Member Names and Lead Name columns in Master_Attendance.xlsx
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_FILE = os.path.join(SCRIPT_DIR, 'Master_Attendance.xlsx')

print("="*80)
print("👁️ FIXING COLUMN VISIBILITY - Master_Attendance.xlsx")
print("="*80)

# Open the workbook
print("\n📖 Opening Master_Attendance.xlsx...")
if not os.path.exists(MASTER_FILE):
    print(f"❌ ERROR: Master_Attendance.xlsx not found at {MASTER_FILE}")
    exit(1)

wb = load_workbook(MASTER_FILE)
print(f"  ✓ Workbook opened successfully")

# Process each sheet
for sheet_name in wb.sheetnames:
    print(f"\n🔧 Processing sheet: '{sheet_name}'")
    ws = wb[sheet_name]
    
    # Check current column properties
    print(f"\n  📊 Current Column Properties:")
    print(f"    Column A (1):")
    print(f"      Hidden: {ws.column_dimensions['A'].hidden}")
    print(f"      Width: {ws.column_dimensions['A'].width}")
    
    print(f"    Column B (2):")
    print(f"      Hidden: {ws.column_dimensions['B'].hidden}")
    print(f"      Width: {ws.column_dimensions['B'].width}")
    
    print(f"    Column C (3):")
    print(f"      Hidden: {ws.column_dimensions['C'].hidden}")
    print(f"      Width: {ws.column_dimensions['C'].width}")
    
    # Fix Column A - Team Member Names
    print(f"\n  ✅ Fixing Column A (Team Member Names):")
    ws.column_dimensions['A'].hidden = False
    ws.column_dimensions['A'].width = 30  # Wide enough to show full names
    print(f"    ✓ Unhidden: True")
    print(f"    ✓ Width set to: 30")
    
    # Fix Column B - Lead Name
    print(f"\n  ✅ Fixing Column B (Lead Name):")
    ws.column_dimensions['B'].hidden = False
    ws.column_dimensions['B'].width = 25  # Wide enough to show lead names
    print(f"    ✓ Unhidden: True")
    print(f"    ✓ Width set to: 25")
    
    # Verify Column C - Location
    print(f"\n  ✅ Verifying Column C (Location):")
    ws.column_dimensions['C'].hidden = False
    ws.column_dimensions['C'].width = 22
    print(f"    ✓ Unhidden: True")
    print(f"    ✓ Width set to: 22")
    
    # Set date columns width
    print(f"\n  ✅ Setting date columns (D onwards) width:")
    for col in range(4, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = 8
    print(f"    ✓ Date columns width set to: 8")
    
    # Verify data is present
    print(f"\n  📋 Verifying data in columns:")
    print(f"    Row 1:")
    print(f"      Column A: '{ws.cell(row=1, column=1).value}'")
    print(f"      Column B: '{ws.cell(row=1, column=2).value}'")
    print(f"      Column C: '{ws.cell(row=1, column=3).value}'")
    
    print(f"\n    Row 4 (First data row):")
    print(f"      Column A: '{ws.cell(row=4, column=1).value}'")
    print(f"      Column B: '{ws.cell(row=4, column=2).value}'")
    print(f"      Column C: '{ws.cell(row=4, column=3).value}'")

# Save the workbook
print(f"\n💾 Saving workbook with visible columns...")
try:
    wb.save(MASTER_FILE)
    print(f"  ✓ File saved successfully!")
except Exception as e:
    print(f"  ❌ ERROR saving file: {str(e)}")
    print(f"     Make sure the file is not open in Excel")
    exit(1)

wb.close()

print("\n" + "="*80)
print("✅ COLUMN VISIBILITY FIXED!")
print("="*80)
print("\nChanges made:")
print("  ✓ Column A (Team Member Names) - Unhidden, Width: 30")
print("  ✓ Column B (Lead Name) - Unhidden, Width: 25")
print("  ✓ Column C (Location) - Unhidden, Width: 22")
print("  ✓ Date columns (D onwards) - Width: 8")
print("\n📝 Next steps:")
print("  1. Close the Excel file if it's currently open")
print("  2. Reopen Master_Attendance.xlsx")
print("  3. Columns A and B should now be clearly visible")
print("\n" + "="*80)
