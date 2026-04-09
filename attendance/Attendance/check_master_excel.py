"""
Diagnostic script to check Master_Attendance.xlsx structure
"""
import os
from openpyxl import load_workbook

MASTER_FILE = os.path.join(os.path.dirname(__file__), 'Master_Attendance.xlsx')

print("="*80)
print("Master Excel File Diagnostics")
print("="*80)

if not os.path.exists(MASTER_FILE):
    print(f"\n❌ File not found: {MASTER_FILE}")
    exit(1)

print(f"\n✓ File exists: {MASTER_FILE}")
print(f"  Size: {os.path.getsize(MASTER_FILE):,} bytes")

try:
    wb = load_workbook(MASTER_FILE, read_only=True, data_only=True)
    print(f"\n✓ File opened successfully")
    print(f"\nSheets in workbook ({len(wb.sheetnames)}):")
    
    for sheet_name in wb.sheetnames:
        print(f"\n  📋 Sheet: '{sheet_name}'")
        sheet = wb[sheet_name]
        print(f"     Rows: {sheet.max_row}, Columns: {sheet.max_column}")
        
        # Show first 10 rows
        print(f"     First 10 rows:")
        for row_idx in range(1, min(sheet.max_row + 1, 11)):
            row_data = []
            for col_idx in range(1, min(sheet.max_column + 1, 8)):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if val:
                    row_data.append(f"{col_idx}:'{val}'")
            if row_data:
                print(f"       Row {row_idx}: {', '.join(row_data)}")
    
    wb.close()
    print(f"\n✓ File closed successfully")
    
except Exception as e:
    print(f"\n❌ Error: {str(e)}")
    import traceback
    traceback.print_exc()

print("\n" + "="*80)
